import cv2
import pytesseract
import os
import re
from openpyxl import Workbook
import logging
import datetime
from concurrent.futures import ThreadPoolExecutor

log_file = "app.log"
logging.basicConfig(filename=log_file, level=logging.DEBUG, format='%(asctime)s %(levelname)s: %(message)s')
logging.info('L\'application a été ouverte le {}'.format(datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S')))
try:
    pytesseract.pytesseract.tesseract_cmd = r'C:\Program Files\Tesseract-OCR\tesseract.exe'
except Exception as e:
    print("Error can't find tesseract:", e)

def apply_threshold(image, threshold_value):
    _, thresholded = cv2.threshold(image, threshold_value, 255, cv2.THRESH_BINARY)
    return thresholded

def remove_lines(image):
    horizontal_kernel = cv2.getStructuringElement(cv2.MORPH_RECT, (24, 1))
    vertical_kernel = cv2.getStructuringElement(cv2.MORPH_RECT, (1, 24))
    horizontal_lines = cv2.morphologyEx(image, cv2.MORPH_OPEN, horizontal_kernel, iterations=2)
    vertical_lines = cv2.morphologyEx(image, cv2.MORPH_OPEN, vertical_kernel, iterations=2)
    lines = cv2.addWeighted(horizontal_lines, 1, vertical_lines, 1, 0)
    return cv2.addWeighted(image, 1, lines, -1, 0)

def extract_data_for_image(image_path):
    try:
        image = cv2.imread(image_path)
        if image.shape[0] > 750 or image.shape[1] > 1624:
            scale_factor = min(750 / image.shape[0], 1624 / image.shape[1])
            image = cv2.resize(image, (0, 0), fx=scale_factor, fy=scale_factor)
        # Coordinates extraction
        x, y, w, h = 440, 690, 305, 38
        crop_img = image[y:y + h, x:x + w]
        gray = cv2.cvtColor(crop_img, cv2.COLOR_BGR2GRAY)
        text1 = pytesseract.image_to_string(gray) #coordonnée
        pattern = r"[-]?\d+[.]\d+,\s*[-]?\d+[.]\d+"
        matches = re.findall(pattern, text1)
        coordinates = matches[0].split(',') if len(matches) > 0 else ['', '']
        # Temperature extraction
        roi = (660, 550, 460, 37)
        x, y, w, h = roi
        if x + w > image.shape[1]:
            w = image.shape[1] - x
        if y + h > image.shape[0]:
            h = image.shape[0] - y
        roi_img = image[y:y + h, x:x + w]
        gray = cv2.cvtColor(roi_img, cv2.COLOR_BGR2GRAY)
        text = pytesseract.image_to_string(gray) #température
        numero_image = os.path.basename(image_path).split('.')[0]
        return (numero_image, text1, text, coordinates[0], coordinates[1])
    except Exception as e:
        logging.error('Erreur lors du traitement de l\'image {}: {}'.format(image_path, str(e)))
        return ('', '', '', '', '')

def extract_data(folder_path, save_path):
    try:
        # Create a new workbook
        wb = Workbook()
        ws = wb.active
        # Add column headers
        headers = ['FEEDER', 'TRONCONS', 'TEMPERATURE DEFAUTS', 'NUMERO D\'IMAGE', 'DEFAULTS', 'LATITUDES', 'LONGITUDES']
        ws.append(headers)
        # Get the list of files in the folder
        files = [os.path.join(folder_path, file) for file in os.listdir(folder_path)]
        with ThreadPoolExecutor() as executor:
            results = list(executor.map(extract_data_for_image, files))
        # Write the data to the Excel file
        for i, (numero_image, text1, text, latitudes, longitudes) in enumerate(results):
            row = i + 2  # Start at row 2 (after the column headers)
            ws.cell(row=row, column=4, value=numero_image)
            ws.cell(row=row, column=3, value=text)
            ws.cell(row=row, column=6, value=latitudes)  # Latitude
            ws.cell(row=row, column=7, value=longitudes)  # Longitude
            # Print the extracted data after saving the Excel file
        print("Extracted data:")
        for i, (numero_image, text1, text, latitudes, longitudes) in enumerate(results):
            print(f'{numero_image}, {text1}, {text}, {latitudes}, {longitudes}')
        # Save the Excel file in the selected folder
        wb.save(save_path)
        logging.info('Extraction terminée')
    except Exception as e:
        logging.error('Erreur : {}'.format(str(e)))
        
if __name__ == "__main__":
    folder_path = r"C:\Users\GAMER\Documents\Projects\dat_photos_thermique"
    save_path = "mon_fichier.xlsx"
    extract_data(folder_path, save_path)
