import os
import openpyxl
from openpyxl.styles import Border, Side
from flask import Flask, flash, render_template, request, Blueprint, url_for, redirect,session
import mysql.connector
from datetime import datetime
import config
import math
from database_connection import execute_query,db_config_client
app5_blueprint = Blueprint('app5', __name__)

@app5_blueprint.route('/resume_rapport_quantification')
def resume_rapport_quantification():
    return render_template('resume_rapport_quantification.html')

@app5_blueprint.route('/resume_quantification_rapport', methods=["POST","GET"])
def resume_quantification_rapport():
    email = session.get('email')
    SERVER_FOLDER = config.SERVER_FOLDER
    REFERENCE_FILE = './Exemple_rapport/Quantification_Statistique_exemple.xlsx'
    date_today = datetime.now().strftime("%Y-%m-%d")
    
    if request.method == 'POST':
        sessionFields = request.form.getlist('session[]')
        usernameFields = request.form.getlist('utilisation[]')
        folder_paths = []
        
        for sessionField, usernameField in zip(sessionFields, usernameFields):
            # Vérifier que l'utilisateur et la session existent dans la base de données
            query = "SELECT * FROM client.user WHERE email = %s"
            user_result = execute_query(query, values=(usernameField,), fetchall=True)
            
            if not user_result:
                flash('Session introuvable dans la base de données pour l\'utilisateur ' + usernameField, 'error')
                return redirect(url_for('app5.resume_rapport_quantification'))
            
            # Vérifier si le dossier de l'utilisateur et de la session existe sur le serveur
            user_folder_path = os.path.join(SERVER_FOLDER, usernameField)
            session_folder_path = os.path.join(user_folder_path, sessionField)
            
            if not os.path.exists(session_folder_path):
                flash("Le dossier de l'utilisateur et de la session spécifiée n'existe pas pour l'utilisateur " + usernameField, 'error')
                return redirect(url_for('app5.resume_rapport_quantification'))
            
            folder_paths.append(session_folder_path)
        
        try:
            # Load the reference file
            wb_reference = openpyxl.load_workbook(REFERENCE_FILE)
            feuille_reference = wb_reference.active
            
            # Define border style
            border_style = Border(left=Side(style='thin'),
                                  right=Side(style='thin'),
                                  top=Side(style='thin'),
                                  bottom=Side(style='thin'))
            
            # Find and process the files in the server folder
            for folder_path in folder_paths:
                for root, dirs, files in os.walk(folder_path):
                    for file in files:
                        if file.endswith(".xlsx") and "Quantification" in file:
                            file_path = os.path.join(root, file)
                            
                            # Load the file
                            try:
                                wb_file = openpyxl.load_workbook(file_path, data_only=True)
                                feuille_file = wb_file.active
                                
                                # Get the data from D7 and D8
                                data_d7 = feuille_file['D7'].value.split(': ')[1]
                                data_d8 = feuille_file['D8'].value.split(': ')[1]
                                
                                # Write the data to the reference sheet
                                feuille_reference['D7'] = feuille_reference['D7'].value + ', ' + data_d7
                                feuille_reference['D8'] = feuille_reference['D8'].value + ', ' + data_d8
                                
                                # Get the data from C11 and D11 onwards
                                for row_num in range(11, feuille_file.max_row + 1):
                                    data_c11 = feuille_file[f'C{row_num}'].value
                                    data_d11 = feuille_file[f'D{row_num}'].value
                                    feuille_reference.append(['', '', data_c11, data_d11])
                                    
                                # Apply border to the added data
                                for row in feuille_reference.iter_rows(
                                        min_row=feuille_reference.max_row - (feuille_file.max_row - 10),
                                        max_row=feuille_reference.max_row,
                                        min_col=3,
                                        max_col=4):
                                    for cell in row:
                                        cell.border = border_style
                            except openpyxl.utils.exceptions.InvalidFileException:
                                flash("Le fichier {} n'est pas un fichier Excel valide.".format(file), 'error')
            
            nom_du_fichier_sortie = "Résumé_final_Quantification_du_{}.xlsx".format(date_today)
            # Save the reference file with the added data
            combined_file_path = os.path.join(SERVER_FOLDER, "Rapport_Finaux")
            
            if not os.path.exists(combined_file_path):
                try:
                    os.makedirs(combined_file_path)
                except OSError as e:
                    flash('Erreur lors de la création du dossier : ' + str(e), 'error')
                    return redirect(url_for('app5.resume_rapport_quantification'))
            
            wb_reference.save(os.path.join(combined_file_path, nom_du_fichier_sortie))
            
            # Calculate file size and type
            file_size_bytes = os.path.getsize(os.path.join(combined_file_path, nom_du_fichier_sortie))
            file_size = convert_size(file_size_bytes)
            file_type = 'excel/xlsx'
            
            # Insert file info into the database
            insert_file_info_to_db(email, nom_du_fichier_sortie, combined_file_path, file_type, date_today, file_size)
            
            flash("Rapport final de quantification généré avec succès.")
            # Proceed to the completed_report page
            return redirect(url_for('app5.resume_rapport_quantification'))
        except FileNotFoundError:
            flash("Le fichier de référence {} est introuvable.".format(REFERENCE_FILE), 'error')
            return redirect(url_for('app5.resume_rapport_quantification'))
    
    return render_template('resume_rapport_quantification.html')

def insert_file_info_to_db(email, file_name, file_path, file_type, creation_date, file_size):
    conn = mysql.connector.connect(**db_config_client)
    cursor = conn.cursor()
    query = "INSERT INTO fichiers (email_utilisateur, nom_fichier, chemin_fichier, type_fichier, date_creation, poids_fichier) " \
            "VALUES (%s, %s, %s, %s, %s, %s)"
    values = (email, file_name, file_path, file_type, creation_date, file_size)
    cursor.execute(query, values)
    conn.commit()
    cursor.close()
    conn.close()

def convert_size(size_bytes):
    if size_bytes == 0:
        return "0B"
    size_name = ("B", "KB", "MB", "GB", "TB", "PB", "EB", "ZB", "YB")
    i = 0 if size_bytes == 0 else int(math.floor(math.log(size_bytes, 1024)))
    p = math.pow(1024, i)
    s = round(size_bytes / p, 2)
    return "%s %s" % (s, size_name[i])
