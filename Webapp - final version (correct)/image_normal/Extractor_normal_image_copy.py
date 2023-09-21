import os
from datetime import datetime
import xlsxwriter
import simplekml
from PIL import Image as IMG
import openpyxl
from openpyxl.styles import Alignment, Font, Border, Side
import json
from flask import  flash,Blueprint,redirect,request,make_response,session
from urllib.parse import unquote
import locale
from image_normal.Generation_doc_word import generate_report_document
from openpyxl.chart import DoughnutChart, Reference
from database_connection import execute_query,db_config_client
import math
import mysql.connector

app2_blueprint = Blueprint('app2', __name__)

@app2_blueprint.route('/generate_report_data', methods=['POST'])
def generate_report_data():
    try:
        troncon = unquote(request.cookies.get('troncon'))
        names = unquote(request.cookies.get('names'))
        date = unquote(request.cookies.get('date'))
        feeder = unquote(request.cookies.get('feeder'))
        zone = unquote(request.cookies.get('zone'))
        groupement = unquote(request.cookies.get('groupement'))
        flash(troncon)
        flash(names)
        flash('ok')
        # Stocker les valeurs dans des cookies
        response = make_response(redirect('/traitement'))
        response.set_cookie('troncon', troncon)
        response.set_cookie('names', names)
        response.set_cookie('date', date)
        response.set_cookie('feeder', feeder)
        response.set_cookie('zone', zone)
        response.set_cookie('groupement', groupement)
        # Votre code pour générer le rapport en utilisant les valeurs récupérées
        return response
    except Exception as e:
        flash("Error occurred:", e)

def generate_report(session_folder_path):
    email = session.get('email')
    date = datetime.now().strftime("%Y-%m-%d")
    troncon = request.cookies.get('troncon')
    feeder = request.cookies.get('feeder')
    groupement = request.cookies.get('groupement')
    # Création du dossier "rapport" s'il n'existe pas
    rapport_folder_path = os.path.join(session_folder_path, "Rapport_defauts_visible")
    if not os.path.exists(rapport_folder_path):
        os.makedirs(rapport_folder_path)
    # Chemin du fichier xlsx
    nom_du_fichier = "Rapport_defaut_visible_goupement_de_troncon_{}_du_{}.xlsx".format(groupement, date)
    xlsx_path = os.path.join(rapport_folder_path,nom_du_fichier)
    # Création du classeur Excel
    workbook = xlsxwriter.Workbook(xlsx_path)
    worksheet = workbook.add_worksheet()
    bold_format = workbook.add_format({'bold': True, 'align': 'center'})
    # Ajout des en-têtes de colonnes
    column_names = ['feeder', 'troncon', 'image', 'défauts', 'latitude', 'longitude', 'date/heure']
    for i, column_name in enumerate(column_names):
        worksheet.write(0, i, column_name, bold_format)
        cell_format = workbook.add_format({'align': 'center'})
        if column_name in ['feeder', 'troncon', 'défauts']:
            worksheet.set_column(i, i, 18, cell_format)
        elif column_name == 'image':
            worksheet.set_column(i, i, 50, cell_format)
        elif column_name == 'date/heure':
            worksheet.set_column(i, i, 20, cell_format)
        else:
            worksheet.set_column(i, i, 25, cell_format)

    # Parcours des images dans le dossier de session
    row_num = 1
    image_data = {}  # Dictionnaire pour stocker les données des images
    for root, dirs, files in os.walk(session_folder_path):
        for filename in files:
            if filename.lower().endswith(('.jpg', '.png', '.jpeg')):
                image_path = os.path.join(root, filename)
                data = image_coordinates(image_path)
                if data:
                    image_name = data[0]
                    parent_dir = os.path.basename(os.path.dirname(image_path))
                    defect = f"{parent_dir}/{data[1]}"
                    latitude = data[2]
                    longitude = data[3]
                    datetime_str = f"{data[4].date()} {data[4].time()}"

                    if image_name in image_data:
                        # Mettre à jour les informations existantes
                        image_data[image_name]['défauts'] += f" / {defect}"
                    else:
                        # Ajouter une nouvelle entrée dans le dictionnaire
                        image_data[image_name] = {
                            'image': image_name,
                            'défauts': defect,
                            'latitude': latitude,
                            'longitude': longitude,
                            'date/heure': datetime_str
                        }
    # Écriture des données dans le fichier Excel à partir du dictionnaire
    for row_num, data in enumerate(image_data.values(), start=1):
        # Ajoutez les valeurs 'feeder' et 'troncon' aux colonnes correspondantes
        worksheet.write(row_num, column_names.index('feeder'), feeder)
        worksheet.write(row_num, column_names.index('troncon'), troncon)
        worksheet.write(row_num, column_names.index('image'), data['image'])
        worksheet.write(row_num, column_names.index('défauts'), data['défauts'])
        worksheet.write(row_num, column_names.index('latitude'), data['latitude'])
        worksheet.write(row_num, column_names.index('longitude'), data['longitude'])
        worksheet.write(row_num, column_names.index('date/heure'), data['date/heure'])

    # Fermeture du classeur Excel
    workbook.close()
    # Générer le fichier KML à partir du fichier XLSX
    generate_kml(xlsx_path, session_folder_path)
    # Générer le rapport "Quanticattion"
    generate_quantification_report(session_folder_path, image_data.copy())
    # Après la génération du rapport "Quanticattion"
    generate_resume_rapport(session_folder_path,image_data.copy())
    generate_report_document(session_folder_path)
    file_size_bytes = os.path.getsize(xlsx_path )
    file_size = convert_size(file_size_bytes)
    creation_date = datetime.now()  
    file_type = 'excel/xlsx'
    insert_file_info_to_db(email,nom_du_fichier,xlsx_path,file_type,creation_date,file_size)

def generate_resume_rapport(session_folder_path,image_data):
    email = session.get('email')
    troncon = request.cookies.get('troncon')
    names = request.cookies.get('names')
    date = request.cookies.get('date')
    feeder = request.cookies.get('feeder')
    zone = request.cookies.get('zone')
    curent_date = datetime.now().strftime("%Y-%m-%d")  # Convert the date to a string with the format "YYYY-MM-DD"

    # Charger les normes_conseils des défauts à partir du fichier JSON
    with open('./phrase_normes_conseils/normes_conseils.json',encoding='utf-8') as f:
        normes_conseils_data = json.load(f)
    # Création du dossier "rapport" s'il n'existe pas
    rapport_folder_path = os.path.join(session_folder_path, "Rapport_defauts_visible")
    if not os.path.exists(rapport_folder_path):
        os.makedirs(rapport_folder_path)

    # Créer une copie du fichier
    wb_copy = openpyxl.load_workbook('./Exemple_rapport/resume_rapport_visibles_exemple.xlsx')
    feuille_copy = wb_copy.active

    # Écrire dans les cellules spécifiques du fichier d'origine
    feuille_copy['A8'] = "Feeder : " + feeder
    feuille_copy['F3'] = "Date : " + date
    feuille_copy['G6'] = "Noms : " + names
    feuille_copy['G8'] = "Zone : " + zone
    # Définir les styles de cellule
    font = Font(name='Calibri', size=18)
    alignment = Alignment(horizontal='center', vertical='center')
    border = Border(top=Side(border_style='thick'),
                    bottom=Side(border_style='thick'),
                    left=Side(border_style='thick'),
                    right=Side(border_style='thick'))
    # Générer des données pour chaque colonne
    row_num = 12
    for image_info in image_data.values():
                # Logique pour les normes_conseils_data I et J
        defauts = image_info['défauts'].split(" / ")  # Divise les défauts par "/"
        colonne_I_values = []
        colonne_J_values = []

        for defaut in defauts:
            if defaut in normes_conseils_data:
                colonne_I_values.append(normes_conseils_data[defaut]['I'])
                colonne_J_values.append(normes_conseils_data[defaut]['J'])

        feuille_copy[f'A{row_num}'] = image_info['date/heure']  # Date/Heure
        feuille_copy[f'B{row_num}'] = feeder #feeder
        feuille_copy[f'C{row_num}'] = troncon #troncon
        feuille_copy[f'D{row_num}'] = ""#longeur
        feuille_copy[f'E{row_num}'] = image_info['image']  # Nom de l'image
        latitude_nom = "Latitude"
        longitude_nom = "Longitude"
        feuille_copy[f'F{row_num}'] = f"{latitude_nom} {float(image_info['latitude']):.8f}, {longitude_nom} {float(image_info['longitude']):.8f}"
        feuille_copy[f'G{row_num}'] = image_info['défauts']  # Défaut de l'image
        feuille_copy[f'H{row_num}'] = ""#urgences
        #Écriture des valeurs dans les colonnes I et J
        feuille_copy[f'I{row_num}'] = '\n'.join(colonne_I_values)
        feuille_copy[f'J{row_num}'] = '\n'.join(colonne_J_values)
        feuille_copy[f'K{row_num}'] = row_num - 11  # Compter de 1 à n

        # Appliquer les styles aux cellules
        for col in ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', 'J', 'K']:
            cell = feuille_copy[f'{col}{row_num}']
            cell.alignment = alignment
            cell.font = font
            cell.border = border

        # Définir la hauteur des lignes
        feuille_copy.row_dimensions[row_num].height = 138

        row_num += 1  # Incrémenter le numéro de ligne
    # Appliquer le style au titre "Urgences"
    cell_title = feuille_copy['H11']
    cell_title.fill = openpyxl.styles.PatternFill(fill_type='solid', fgColor='FF0000')  # Rouge
    cell_title.font = Font(color="FFFFFF", bold=True)  # Texte blanc en gras
    nom_du_fichier = "Résumé_des_rapport_visibles_du_feeder_{}_du_{}.xlsx".format(feeder,curent_date)
    resume_rapport_path = os.path.join(rapport_folder_path,nom_du_fichier )  # Retourner chemin du fichier
    # Enregistrer les modifications dans le fichier d'origine
    wb_copy.save(resume_rapport_path)
    file_size_bytes = os.path.getsize(resume_rapport_path)
    file_size = convert_size(file_size_bytes)
    creation_date = datetime.now()  
    file_type = 'excel/xlsx'
    insert_file_info_to_db(email,nom_du_fichier,resume_rapport_path,file_type,creation_date,file_size)

    
def image_coordinates(image_path):
    with open(image_path, 'rb') as src:
        img = IMG.open(src)
        exif_data = img._getexif()
        if exif_data:
            try:
                gps_info = exif_data[34853]
                latitude = decimal_coords(gps_info[2], gps_info[1])
                longitude = decimal_coords(gps_info[4], gps_info[3])
            except KeyError:
                flash(f'{image_path}: Pas de coordonnées')
                return
        else:
            flash(f'{image_path}: L\'image ne contient pas d\'informations EXIF')
            return
        #print(f"{image_path}, Version du système d'exploitation : {exif_data.get(305, 'Inconnue')}")
        # Conversion de la date et de l'heure en objet datetime
        datetime_str = exif_data.get(306)
        datetime_obj = datetime.strptime(datetime_str, '%Y:%m:%d %H:%M:%S')
        # Retour des données pour l'écriture dans le fichier Excel
        filename = os.path.splitext(os.path.basename(image_path))[0]
        return filename, '', latitude, longitude, datetime_obj

def decimal_coords(coords, ref):
    decimal_degrees = coords[0] + coords[1] / 60 + coords[2] / 3600
    if ref == "S" or ref == "W":
        decimal_degrees = -decimal_degrees
    return decimal_degrees

def generate_kml(xlsx_path, session_folder_path):
    # Définir la configuration régionale pour utiliser une virgule en tant que séparateur décimal
    locale.setlocale(locale.LC_NUMERIC, 'fr_FR')
    # Création du dossier "Data_kml" s'il n'existe pas
    data_folder_path = os.path.join(session_folder_path, "Data_kml")
    if not os.path.exists(data_folder_path):
        os.makedirs(data_folder_path)
    # Charger le fichier XLSX
    workbook = openpyxl.load_workbook(xlsx_path)
    worksheet = workbook.active
    # Parcourir les lignes du fichier XLSX
    for row in worksheet.iter_rows(min_row=2, values_only=True):  # Ignorer la première ligne d'en-tête
        feeder = row[0]
        troncon = row[1]
        filename = row[2]
        defect = row[3]
        latitude_value = row[4]
        longitude_value = row[5]
        date_heure = row[6]
        # Vérifier si les valeurs de latitude et de longitude sont déjà des nombres (float)
        if isinstance(latitude_value, float):
            latitude = latitude_value
        else:
            # Vérifier et convertir la valeur de latitude
            try:
                latitude = locale.atof(latitude_value)
            except ValueError:
                flash(f"{filename}: Latitude invalide")
                continue
        if isinstance(longitude_value, float):
            longitude = longitude_value
        else:
            # Vérifier et convertir la valeur de longitude
            try:
                longitude = locale.atof(longitude_value)
            except ValueError:
                flash(f"{filename}: Longitude invalide")
                continue
        # Créer un objet KML
        kml = simplekml.Kml()
        point = kml.newpoint()
        # Définir les attributs du point
        point.coords = [(longitude, latitude)]
        point.description = f"Feeder: {feeder}\nTroncon: {troncon}\nDéfaut: {defect}\nNom du fichier: {filename}\nDate et heure: {date_heure}"

        # Enregistrer le fichier KML
        kml_path = os.path.join(data_folder_path, f"{filename}.kml")
        kml.save(kml_path)
    # Fermeture du classeur Excel
    workbook.close()

def generate_quantification_report(session_folder_path, image_data):
    email = session.get('email')
    names = request.cookies.get('names')
    date = request.cookies.get('date')
    feeder = request.cookies.get('feeder')
    rapport_folder_path = os.path.join(session_folder_path, "Rapport_defauts_visible")
    if not os.path.exists(rapport_folder_path):
        os.makedirs(rapport_folder_path)
    curent_date = datetime.now().strftime("%Y-%m-%d")  # Convert the date to a string with the format "YYYY-MM-DD"

    # Write data to the copy sheet
    wb_copy = openpyxl.load_workbook("./Exemple_rapport/Quantification_Statistique_exemple.xlsx")
    feuille_copy = wb_copy.active
    feuille_copy['D7'] = "Noms : " + names
    feuille_copy['D8'] = "Date : " + date
    feuille_copy['D11'] = feeder

    # Merge cells C11, C12, C13 and D11, D12, D13 and E11, E12, E13
    feuille_copy.merge_cells('C11:C13')
    feuille_copy.merge_cells('D11:D13')
    feuille_copy.merge_cells('E11:E13')

    # Define styles and fonts
    font = Font(name="Calibri", size=12)
    alignment = Alignment(horizontal="center", vertical="center")
    thin_border = Border(left=Side(style='thin'),
                         right=Side(style='thin'),
                         top=Side(style='thin'),
                         bottom=Side(style='thin'))

    # Apply styles and borders to specific cells on the copy sheet
    for col in ['A', 'B', 'C', 'D', 'E']:
        for row_num in range(11, feuille_copy.max_row + 1):
            cell = feuille_copy[f'{col}{row_num}']
            cell.alignment = alignment
            cell.font = font
            cell.border = thin_border

    # Add data for each column in the copy sheet starting from A12
    row_num = 12
    defect_count = {}
    for data in image_data.values():
        defects = data['défauts'].split(' / ')
        for defect in defects:
            if defect in defect_count:
                defect_count[defect] += 1
            else:
                defect_count[defect] = 1

    for defect, count in defect_count.items():
        feuille_copy.append(['', '', defect, count])

    # Total of defects in green cell
    feuille_copy.append(['', '', 'Totale des défauts', sum(defect_count.values())])
    total_cell = feuille_copy.cell(row=feuille_copy.max_row, column=4)
    total_cell.font = Font(name="Calibri", size=12, color="00AA00")  # Green font color

    # Merge cell containing "DRS" with cells below
    drs_cell = feuille_copy['A12']
    feuille_copy.merge_cells(start_row=12, start_column=1, end_row=feuille_copy.max_row, end_column=1)
    drs_cell.alignment = alignment

    # Add Pie Chart at the end
    pie = DoughnutChart()
    labels = Reference(feuille_copy, min_col=3, min_row=14, max_row=feuille_copy.max_row - 1)
    data = Reference(feuille_copy, min_col=4, min_row=11, max_row=feuille_copy.max_row - 1)
    pie.add_data(data, titles_from_data=True)
    pie.set_categories(labels)
    chart_cell = feuille_copy.cell(row=feuille_copy.max_row + 2, column=1)
    feuille_copy.add_chart(pie, chart_cell.coordinate)
    nom_du_fichier = "Quantification_du_feeder_{}_du_{}.xlsx".format(feeder, curent_date)
    # Save the copy
    quantification_folder_path = os.path.join(rapport_folder_path, nom_du_fichier)
    wb_copy.save(quantification_folder_path)
    file_size_bytes = os.path.getsize(quantification_folder_path)
    file_size = convert_size(file_size_bytes)
    creation_date = datetime.now()  
    file_type = 'excel/xlsx'
    insert_file_info_to_db(email,nom_du_fichier,quantification_folder_path,file_type,creation_date,file_size)

def insert_file_info_to_db(email, file_name, file_path, file_type, creation_date, file_size):
    conn = mysql.connector.connect(**db_config_client)
    cursor = conn.cursor()
    query = "INSERT INTO fichiers (email_utilisateur, nom_fichier, chemin_fichier, type_fichier, date_creation, poids_fichier) " \
            "VALUES (%s, %s, %s, %s, %s, %s)"
    values = (email, file_name, file_path, file_type, creation_date, file_size)
    cursor.execute(query, values)
    conn.commit()
    cursor.close()
    conn.close()

def convert_size(size_bytes):
    if size_bytes == 0:
        return "0B"
    size_name = ("B", "KB", "MB", "GB", "TB", "PB", "EB", "ZB", "YB")
    i = 0 if size_bytes == 0 else int(math.floor(math.log(size_bytes, 1024)))
    p = math.pow(1024, i)
    s = round(size_bytes / p, 2)
    return "%s %s" % (s, size_name[i])