import os
import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side
from flask import Flask, flash, render_template, request, Blueprint, url_for, redirect,session
import mysql.connector
from datetime import datetime
import config
import math

app4_blueprint = Blueprint('app4', __name__)
# Configuration de la base de données client
db_config_client = {
    'host': config.DB_HOST,
    'user': config.DB_USER,
    'password': config.DB_PASSWORD,
    'database': config.DB_DATABASE
}

# Fonction utilitaire pour la base de données client
def execute_query(query, values=None, fetchall=False):
    try:
        conn = mysql.connector.connect(**db_config_client)
        cursor = conn.cursor(buffered=True)
        cursor.execute(query, values)
        conn.commit()
        result = cursor.fetchall() if fetchall else None
        cursor.close()
        conn.close()
        return result
    except mysql.connector.Error as error:
        flash('Erreur lors de l\'exécution de la requête : ' + str(error), 'error')
        return None

@app4_blueprint.route('/resume_rapport')
def v():
    return render_template('resume_rapport.html')

@app4_blueprint.route('/resume_rapport', methods=['POST'])
def resume_rapport():
    email = session.get('email')
    SERVER_FOLDER = config.SERVER_FOLDER
    INPUT_FILE = './Exemple_rapport/resume_rapport_visibles_exemple.xlsx'
    date_today = datetime.now().strftime("%Y-%m-%d")
    
    if request.method == 'POST':
        sessionFields = request.form.getlist('session[]')
        usernameFields = request.form.getlist('utilisation[]')
        folder_paths = []  # Liste pour stocker les chemins des dossiers de session et d'utilisateur
        
        for sessionField, usernameField in zip(sessionFields, usernameFields):
            # Vérifiez que l'utilisateur et la session existent dans la base de données
            query = "SELECT * FROM client.user WHERE email = %s"
            user_result = execute_query(query, values=(usernameField,), fetchall=True)
            
            if not user_result:
                flash('Session introuvable dans la base de données pour l\'utilisateur ' + usernameField, 'error')
                return redirect(url_for('app4.resume_rapport'))
            
            # Vérifiez si le dossier de l'utilisateur et de la session existe sur le serveur
            user_folder_path = os.path.join(SERVER_FOLDER, usernameField)
            session_folder_path = os.path.join(user_folder_path, sessionField)
            
            if not os.path.exists(session_folder_path):
                flash("Le dossier de l'utilisateur et de la session spécifiée n'existe pas pour l'utilisateur " + usernameField, 'error')
                return redirect(url_for('app4.resume_rapport'))
            
            folder_paths.append(session_folder_path)  # Ajouter le chemin du dossier à la liste
        
        try:
            wb_copy = openpyxl.load_workbook(INPUT_FILE)
            feuille_copy = wb_copy.active
            
            # Définir les styles de cellule
            font = Font(name='Calibri', size=18)
            alignment = Alignment(horizontal='center', vertical='center')
            border = Border(top=Side(border_style='thick'),
                            bottom=Side(border_style='thick'),
                            left=Side(border_style='thick'),
                            right=Side(border_style='thick'))
            
            feuille_copy['A12'].font = font
            feuille_copy['A12'].alignment = alignment
            feuille_copy['A12'].border = border
            
            # Créer une liste pour stocker les noms de fichier
            noms_fichiers = []
            row_count = 12  # On commence à la ligne 13 (après le titre)
            
            for folder_path in folder_paths:
                for foldername, subfolders, filenames in os.walk(folder_path):
                    for filename in filenames:
                        if filename.endswith('.xlsx') and 'Résumé_des_rapport_visibles' in filename:
                            file_path = os.path.join(foldername, filename)
                            
                            try:
                                wb = openpyxl.load_workbook(file_path, read_only=True)
                                feuille = wb.active
                                
                                # Récupérer le nom du fichier à partir de la cellule A8
                                nom_fichier_cell = feuille['A8'].value
                                nom_fichier = nom_fichier_cell.split(':')[-1].strip()
                                
                                # Ajouter le nom du fichier à la liste
                                noms_fichiers.append(nom_fichier)
                                
                                # Parcourir les lignes à partir de la ligne 12
                                for row in feuille.iter_rows(min_row=12, values_only=True):
                                    for col_idx, cell_value in enumerate(row, start=1):
                                        feuille_copy.cell(row=row_count, column=col_idx, value=cell_value)
                                        feuille_copy.cell(row=row_count, column=col_idx).border = border
                                        feuille_copy.cell(row=row_count, column=col_idx).font = font
                                        feuille_copy.cell(row=row_count, column=col_idx).alignment = alignment
                                        # Définir la hauteur des lignes
                                        feuille_copy.row_dimensions[row_count].height = 138
                                    row_count += 1
                            except openpyxl.utils.exceptions.InvalidFileException:
                                flash("Le fichier {} n'est pas un fichier Excel valide.".format(filename), 'error')
            
            # Concaténer les noms de fichier dans le nom de fichier de sortie
            noms_fichiers_str = "_".join(noms_fichiers)
            nom_fichier_sortie = "Résumé_des_rapport_finaux_des_feeder_du_{}_{}.xlsx".format(noms_fichiers_str, date_today)
            regenerate_resume_rapport = os.path.join(SERVER_FOLDER, "Rapport_Finaux")
            
            if not os.path.exists(regenerate_resume_rapport):
                try:
                    os.makedirs(regenerate_resume_rapport)
                except OSError as e:
                    flash("Erreur lors de la création du dossier : " + str(e), 'error')
                    return redirect(url_for('app4.resume_rapport'))
            
            wb_copy.save(os.path.join(regenerate_resume_rapport, nom_fichier_sortie))
            file_size_bytes = os.path.getsize(os.path.join(regenerate_resume_rapport, nom_fichier_sortie))
            file_size = convert_size(file_size_bytes)
            file_type = 'excel/xlsx'
            creation_date = datetime.now()  # Correction : Utilisez datetime.now() sans les parenthèses
            
            insert_file_info_to_db(email, nom_fichier_sortie, os.path.join(email, regenerate_resume_rapport, nom_fichier_sortie), file_type, creation_date, file_size)
            flash("Rapport final généré avec succès.")
            return redirect(url_for('app4.resume_rapport'))
        except FileNotFoundError:
            flash("Le fichier d'entrée {} est introuvable.".format(INPUT_FILE), 'error')
            return redirect(url_for('app4.resume_rapport'))
    return render_template('resume_rapport.html')

def insert_file_info_to_db(email, file_name, file_path, file_type, creation_date, file_size):
    conn = mysql.connector.connect(**db_config_client)
    cursor = conn.cursor()
    query = "INSERT INTO fichiers (email_utilisateur, nom_fichier, chemin_fichier, type_fichier, date_creation, poids_fichier) " \
            "VALUES (%s, %s, %s, %s, %s, %s)"
    values = (email, file_name, file_path, file_type, creation_date, file_size)
    cursor.execute(query, values)
    conn.commit()
    cursor.close()
    conn.close()


def convert_size(size_bytes):
    if size_bytes == 0:
        return "0B"
    size_name = ("B", "KB", "MB", "GB", "TB", "PB", "EB", "ZB", "YB")
    i = int(math.floor(math.log(size_bytes, 1024)))
    p = math.pow(1024, i)
    s = round(size_bytes / p, 2)
    return "%s %s" % (s, size_name[i])
